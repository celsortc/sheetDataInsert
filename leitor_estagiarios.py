"""
Processador de Contratos de Estágio
Versão 2.0 — com pré-visualização, duplicatas, histórico, configurações e mais.
"""

import pdfplumber
import re
import json
import os
import threading
import subprocess
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog, ttk

# ══════════════════════════════════════════════════════════════
#  CONSTANTES E CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════

ORDEM_COLUNAS = [
    "data-inicial", "data-final", "vazio", "nome", "CPF", "carga-horaria", 
    "hr-entrada-saida", "ano/curso", "supervisor", "telefone"
]

PADROES_PADRAO = {
    "data-inicial": r"Vigência de:\s*(.*?)\sAté",
    "data-final":   r"até\s*(.*)",
    "nome":         r"Nome:\s*(.*?)\s+Código",
    "CPF":          r"CPF/MF:\s*(.*)",
    "ano/curso":    r"Regularmente Matriculado:\s*(\d+)",
    "supervisor":   r"Supervisor:\s*(.*?)\sCargo",
    "hr-entrada-saida": r"Horário das\s*(\d{2}:\d{2})\s*as\s*(\d{2}:\d{2})",
    "telefone":     r"Fone:\s*([^\n]+)",
}

# Pasta de configurações do app (AppData no Windows)
APP_CONFIG_DIR = os.path.join(os.path.expanduser("~"), "AppData", "Local", "ProcessadorEstagiarios")
os.makedirs(APP_CONFIG_DIR, exist_ok=True)
CONFIG_FILE   = os.path.join(APP_CONFIG_DIR, "config.json")
HISTORICO_FILE = os.path.join(APP_CONFIG_DIR, "historico.json")

# Cores
VERDE    = "#22c55e"
AMARELO  = "#f59e0b"
VERMELHO = "#ef4444"
AZUL     = "#3b82f6"
ROXO     = "#8b5cf6"
CINZA_BG    = "#1e1e2e"
CINZA_CARD  = "#2a2a3e"
CINZA_BORDA = "#3f3f5c"
BRANCO   = "#e2e8f0"
TEXTO_SUB = "#94a3b8"

ICONES = {"processando": "⏳", "ok": "✅", "erro": "❌", "pendente": "⏸", "pulado": "⏭", "revisao": "✏️"}


# ══════════════════════════════════════════════════════════════
#  CONFIGURAÇÃO PERSISTENTE
# ══════════════════════════════════════════════════════════════

def carregar_config():
    padrao = {
        "caminho_excel": os.path.join(os.path.expanduser("~"), "Documents", "estags.xlsx"),
        "padroes": PADROES_PADRAO.copy()
    }
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # garante que campos novos existam
            for k, v in padrao.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
        except Exception:
            pass
    return padrao

def salvar_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def carregar_historico():
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def salvar_historico(historico):
    with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
        json.dump(historico[-50:], f, ensure_ascii=False, indent=2)  # mantém últimas 50 execuções


# ══════════════════════════════════════════════════════════════
#  CONTROLE DE PROCESSADOS (por pasta)
# ══════════════════════════════════════════════════════════════

def caminho_processados(pasta):
    return os.path.join(pasta, ".processados.json")

def carregar_processados(pasta):
    p = caminho_processados(pasta)
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()

def salvar_processados(pasta, conjunto):
    with open(caminho_processados(pasta), "w", encoding="utf-8") as f:
        json.dump(sorted(conjunto), f, ensure_ascii=False, indent=2)


# ══════════════════════════════════════════════════════════════
#  LÓGICA DE EXTRAÇÃO
# ══════════════════════════════════════════════════════════════

def formatar_nome(nome):
    if not nome:
        return nome
    minusculas = {"da", "de", "do", "das", "dos"}
    palavras = nome.lower().split()
    return " ".join(
        p if (p in minusculas and i != 0) else p.capitalize()
        for i, p in enumerate(palavras)
    )

def formatar_numero(num):
    if not num:
        return num
    limpo = re.sub(r'\D', '', str(num))
    return int(limpo) if limpo else None

def pegar_numero_contrato(nome_arquivo):
    return nome_arquivo.split(' - ')[0]

def extrair_dados_pdf(caminho_pdf, padroes=None):
    """Lê um PDF e retorna (dados_dict, campos_faltantes)."""
    if padroes is None:
        padroes = PADROES_PADRAO

    with pdfplumber.open(caminho_pdf) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)

    dados = {}
    campos_faltantes = []

    campos_simples = ["data-inicial", "data-final", "nome", "CPF", "ano/curso", "supervisor"]
    for campo in campos_simples:
        padrao = padroes.get(campo, PADROES_PADRAO.get(campo, ""))
        if padrao:
            m = re.search(padrao, texto, re.IGNORECASE)
            dados[campo] = m.group(1).strip() if m else None
        else:
            dados[campo] = None
        if not dados[campo]:
            campos_faltantes.append(campo)

    # Horário e carga
    padrao_hr = padroes.get("hr-entrada-saida", PADROES_PADRAO["hr-entrada-saida"])
    m_hr = re.search(padrao_hr, texto, re.IGNORECASE)
    if m_hr:
        entrada, saida = m_hr.group(1), m_hr.group(2)
        dados["hr-entrada-saida"] = f"{entrada} - {saida}"
        dt_e = datetime.strptime(entrada, "%H:%M")
        dt_s = datetime.strptime(saida, "%H:%M")
        dados["carga-horaria"] = int((dt_s - dt_e).total_seconds() / 3600)
    else:
        dados["hr-entrada-saida"] = None
        dados["carga-horaria"] = None
        campos_faltantes.extend(["hr-entrada-saida", "carga-horaria"])

    # Telefone (3º Fone)
    padrao_fone = padroes.get("telefone", PADROES_PADRAO["telefone"])
    fones = re.findall(padrao_fone, texto)
    dados["telefone"] = fones[2].strip() if len(fones) >= 3 else None
    if not dados["telefone"]:
        campos_faltantes.append("telefone")

    # Formatações
    dados["nome"]      = formatar_nome(dados.get("nome"))
    dados["CPF"]       = formatar_numero(dados.get("CPF"))
    dados["telefone"]  = formatar_numero(dados.get("telefone"))
    dados["ano/curso"] = formatar_numero(dados.get("ano/curso"))

    return dados, campos_faltantes


def carregar_contratos_existentes(caminho_excel):
    """Retorna set de números de contrato já na planilha (coluna 'vazio')."""
    existentes = set()
    if not os.path.exists(caminho_excel):
        return existentes
    try:
        wb = load_workbook(caminho_excel)
        ws = wb.active
        idx_vazio = ORDEM_COLUNAS.index("vazio") + 1  # 1-based
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[idx_vazio - 1]
            if val is not None:
                existentes.add(str(val).strip())
    except Exception:
        pass
    return existentes


def ler_pdfs_para_preview(pasta, padroes, callback_progresso):
    """
    Lê todos os PDFs e retorna lista de dicts para pré-visualização.
    Não salva nada. Marca duplicatas e campos faltantes.
    """
    cfg = carregar_config()
    caminho_excel = cfg["caminho_excel"]
    existentes = carregar_contratos_existentes(caminho_excel)
    ja_processados = carregar_processados(pasta)

    arquivos = sorted(f for f in os.listdir(pasta) if f.lower().endswith(".pdf"))
    resultados = []

    for i, arquivo in enumerate(arquivos, 1):
        callback_progresso(i, len(arquivos), arquivo)
        caminho_pdf = os.path.join(pasta, arquivo)
        nc = pegar_numero_contrato(arquivo)

        entry = {
            "arquivo": arquivo,
            "nc": nc,
            "dados": {},
            "campos_faltantes": [],
            "duplicata": nc in existentes,
            "ja_processado": arquivo in ja_processados,
            "erro": None,
            "incluir": True,  # usuário pode desmarcar
        }

        if entry["duplicata"] or entry["ja_processado"]:
            entry["incluir"] = False

        try:
            dados, faltantes = extrair_dados_pdf(caminho_pdf, padroes)
            dados["vazio"] = nc
            entry["dados"] = dados
            entry["campos_faltantes"] = faltantes
        except Exception as e:
            entry["erro"] = str(e)
            entry["incluir"] = False

        resultados.append(entry)

    return resultados


def salvar_resultados(resultados, pasta):
    """Salva no Excel apenas os itens marcados como incluir=True."""
    cfg = carregar_config()
    caminho_excel = cfg["caminho_excel"]

    os.makedirs(os.path.dirname(caminho_excel), exist_ok=True)

    if os.path.exists(caminho_excel):
        wb = load_workbook(caminho_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Estags"
        ws.append(ORDEM_COLUNAS)

    ja_processados = carregar_processados(pasta)
    ok_count = 0

    for entry in resultados:
        if not entry["incluir"] or entry["erro"]:
            continue
        linha = [entry["dados"].get(col) for col in ORDEM_COLUNAS]
        ws.append(linha)
        ja_processados.add(entry["arquivo"])
        ok_count += 1

    wb.save(caminho_excel)
    salvar_processados(pasta, ja_processados)

    # Histórico
    historico = carregar_historico()
    historico.append({
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "pasta": pasta,
        "salvos": ok_count,
        "total": len(resultados),
    })
    salvar_historico(historico)

    return ok_count


def exportar_relatorio_erros(resultados, pasta):
    erros = [e for e in resultados if e["erro"]]
    faltantes = [e for e in resultados if e["campos_faltantes"] and not e["erro"]]
    pulados = [e for e in resultados if e["duplicata"] or e["ja_processado"]]

    if not erros and not faltantes:
        return None

    caminho = os.path.join(pasta, f"relatorio_erros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(f"Relatório de Processamento — {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        f.write("=" * 60 + "\n\n")

        if erros:
            f.write(f"ERROS ({len(erros)} arquivo(s)):\n")
            f.write("-" * 40 + "\n")
            for e in erros:
                f.write(f"  • {e['arquivo']}\n    Erro: {e['erro']}\n\n")

        if faltantes:
            f.write(f"\nCAMPOS NÃO ENCONTRADOS ({len(faltantes)} arquivo(s)):\n")
            f.write("-" * 40 + "\n")
            for e in faltantes:
                f.write(f"  • {e['arquivo']}\n    Faltam: {', '.join(e['campos_faltantes'])}\n\n")

        if pulados:
            f.write(f"\nPULADOS — já processados ou duplicatas ({len(pulados)}):\n")
            f.write("-" * 40 + "\n")
            for e in pulados:
                motivo = "duplicata na planilha" if e["duplicata"] else "já processado antes"
                f.write(f"  • {e['arquivo']} ({motivo})\n")

    return caminho


def limpar_excel():
    cfg = carregar_config()
    caminho = cfg["caminho_excel"]
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estags"
    ws.append(ORDEM_COLUNAS)
    wb.save(caminho)


# ══════════════════════════════════════════════════════════════
#  JANELA DE PRÉ-VISUALIZAÇÃO
# ══════════════════════════════════════════════════════════════

class JanelaPreview(tk.Toplevel):
    """
    Mostra todos os PDFs lidos com seus dados.
    Campos faltantes ficam em vermelho e editáveis.
    Duplicatas aparecem marcadas mas desmarcadas.
    O usuário confirma o que quer salvar.
    """
    def __init__(self, parent, resultados, pasta, callback_salvo):
        super().__init__(parent)
        self.title("Pré-visualização — Confirmar dados antes de salvar")
        self.geometry("1100x620")
        self.configure(bg=CINZA_BG)
        self.resizable(True, True)
        self.grab_set()  # modal

        self._resultados = resultados
        self._pasta = pasta
        self._callback_salvo = callback_salvo
        self._linha_selecionada = None
        self._edits = {}  # iid → {campo: StringVar}

        self._construir_ui()
        self._popular_tabela()

    def _construir_ui(self):
        # Título
        tk.Label(self, text="📋  Revise os dados antes de salvar no Excel",
                 font=("Segoe UI", 12, "bold"), bg=CINZA_BG, fg=BRANCO
                 ).pack(padx=20, pady=(14, 4), anchor="w")

        tk.Label(self,
                 text="✅ verde = OK  |  🟡 amarelo = campos faltantes  |  🔴 vermelho = erro  |  ⏭ cinza = será pulado",
                 font=("Segoe UI", 8), bg=CINZA_BG, fg=TEXTO_SUB
                 ).pack(padx=20, anchor="w")

        # Frame principal com tabela + painel de edição
        frame_main = tk.Frame(self, bg=CINZA_BG)
        frame_main.pack(fill="both", expand=True, padx=20, pady=8)

        # Tabela esquerda
        frame_tabela = tk.Frame(frame_main, bg=CINZA_CARD,
                                highlightthickness=1, highlightbackground=CINZA_BORDA)
        frame_tabela.pack(side="left", fill="both", expand=True)

        cols = ("inc", "arquivo", "nome", "CPF", "status")
        self._tree = ttk.Treeview(frame_tabela, columns=cols, show="headings",
                                  selectmode="browse", height=18)
        self._tree.heading("inc",     text="✔")
        self._tree.heading("arquivo", text="Arquivo")
        self._tree.heading("nome",    text="Nome")
        self._tree.heading("CPF",     text="CPF")
        self._tree.heading("status",  text="Status")
        self._tree.column("inc",     width=28, minwidth=28, anchor="center", stretch=False)
        self._tree.column("arquivo", width=220, minwidth=120)
        self._tree.column("nome",    width=200, minwidth=100)
        self._tree.column("CPF",     width=110, minwidth=80, anchor="center")
        self._tree.column("status",  width=140, minwidth=80, anchor="center")

        style = ttk.Style(self)
        style.configure("Preview.Treeview",
                        background=CINZA_CARD, foreground=BRANCO,
                        fieldbackground=CINZA_CARD, rowheight=26,
                        font=("Segoe UI", 9))
        style.configure("Preview.Treeview.Heading",
                        background=CINZA_BG, foreground=TEXTO_SUB,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Preview.Treeview",
                  background=[("selected", "#3b3b5c")])
        self._tree.configure(style="Preview.Treeview")

        self._tree.tag_configure("ok",       foreground=VERDE)
        self._tree.tag_configure("faltante", foreground=AMARELO)
        self._tree.tag_configure("erro",     foreground=VERMELHO)
        self._tree.tag_configure("pulado",   foreground=TEXTO_SUB)

        scroll = ttk.Scrollbar(frame_tabela, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscroll=scroll.set)
        scroll.pack(side="right", fill="y")
        self._tree.pack(fill="both", expand=True)
        self._tree.bind("<<TreeviewSelect>>", self._ao_selecionar)
        self._tree.bind("<Button-1>", self._ao_clicar)

        # Painel de edição direito
        self._frame_edit = tk.Frame(frame_main, bg=CINZA_CARD, width=280,
                                    highlightthickness=1, highlightbackground=CINZA_BORDA)
        self._frame_edit.pack(side="right", fill="y", padx=(8, 0))
        self._frame_edit.pack_propagate(False)

        tk.Label(self._frame_edit, text="Editar campos",
                 font=("Segoe UI", 10, "bold"), bg=CINZA_CARD, fg=BRANCO
                 ).pack(padx=12, pady=(12, 4), anchor="w")

        self._lbl_arquivo_edit = tk.Label(self._frame_edit, text="Selecione um arquivo",
                                          font=("Segoe UI", 8), bg=CINZA_CARD, fg=TEXTO_SUB,
                                          wraplength=250, justify="left")
        self._lbl_arquivo_edit.pack(padx=12, anchor="w")

        self._frame_campos = tk.Frame(self._frame_edit, bg=CINZA_CARD)
        self._frame_campos.pack(fill="both", expand=True, padx=12, pady=8)

        tk.Button(self._frame_edit, text="💾 Aplicar edições",
                  command=self._aplicar_edicoes,
                  bg=AZUL, fg="white", relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=10, pady=5, cursor="hand2"
                  ).pack(padx=12, pady=8, fill="x")

        # Rodapé
        frame_bot = tk.Frame(self, bg=CINZA_BG)
        frame_bot.pack(fill="x", padx=20, pady=(0, 14))

        self._lbl_resumo = tk.Label(frame_bot, text="",
                                    font=("Segoe UI", 9), bg=CINZA_BG, fg=TEXTO_SUB)
        self._lbl_resumo.pack(side="left")

        tk.Button(frame_bot, text="✖  Cancelar",
                  command=self.destroy,
                  bg=CINZA_CARD, fg=VERMELHO, relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=14, pady=6, cursor="hand2"
                  ).pack(side="right", padx=(6, 0))

        self._btn_salvar = tk.Button(frame_bot, text="💾  Salvar selecionados no Excel",
                  command=self._confirmar_salvar,
                  bg=VERDE, fg="#0f172a", relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=14, pady=6, cursor="hand2"
                  )
        self._btn_salvar.pack(side="right")

    def _popular_tabela(self):
        for row in self._tree.get_children():
            self._tree.delete(row)

        incluidos = 0
        for entry in self._resultados:
            arq = entry["arquivo"]
            nome = entry["dados"].get("nome") or "—"
            cpf  = str(entry["dados"].get("CPF") or "—")

            if entry["erro"]:
                status = "❌ Erro"
                tag = "erro"
                inc_str = "✖"
            elif entry["duplicata"]:
                status = "⏭ Duplicata"
                tag = "pulado"
                inc_str = "—"
            elif entry["ja_processado"]:
                status = "⏭ Já processado"
                tag = "pulado"
                inc_str = "✔" if entry["incluir"] else "☐"
            elif entry["campos_faltantes"]:
                status = f"⚠ Falta: {', '.join(entry['campos_faltantes'][:2])}"
                tag = "faltante"
                inc_str = "✔" if entry["incluir"] else "☐"
            else:
                status = "✅ OK"
                tag = "ok"
                inc_str = "✔"

            if entry["incluir"]:
                incluidos += 1

            self._tree.insert("", "end", iid=arq,
                              values=(inc_str, arq, nome, cpf, status),
                              tags=(tag,))

        total = len(self._resultados)
        self._lbl_resumo.config(
            text=f"{incluidos} de {total} serão salvos"
        )

    def _ao_clicar(self, event):
        """Alterna incluir ao clicar na coluna ✔.
        Duplicatas (mesmo número na planilha) ficam sempre bloqueadas.
        Itens 'já processados' podem ser forçados pelo usuário."""
        region = self._tree.identify_region(event.x, event.y)
        col    = self._tree.identify_column(event.x)
        iid    = self._tree.identify_row(event.y)
        if region == "cell" and col == "#1" and iid:
            entry = next((e for e in self._resultados if e["arquivo"] == iid), None)
            if entry and not entry["erro"] and not entry["duplicata"]:
                entry["incluir"] = not entry["incluir"]
                self._popular_tabela()
                if self._tree.exists(iid):
                    self._tree.selection_set(iid)

    def _ao_selecionar(self, event):
        selecionados = self._tree.selection()
        if not selecionados:
            return
        iid = selecionados[0]
        entry = next((e for e in self._resultados if e["arquivo"] == iid), None)
        if not entry:
            return
        self._linha_selecionada = entry
        self._montar_campos_edicao(entry)

    def _montar_campos_edicao(self, entry):
        for widget in self._frame_campos.winfo_children():
            widget.destroy()

        self._lbl_arquivo_edit.config(text=entry["arquivo"])
        self._edits_vars = {}

        # Todos os campos editáveis exceto "vazio" (vem do nome do arquivo)
        campos_editaveis = [c for c in ORDEM_COLUNAS if c != "vazio"]

        for campo in campos_editaveis:
            val = entry["dados"].get(campo)
            cor_label = VERMELHO if campo in entry["campos_faltantes"] else TEXTO_SUB

            # Frame de label (para carga-horaria mostrar dica ao lado)
            frame_label = tk.Frame(self._frame_campos, bg=CINZA_CARD)
            frame_label.pack(fill="x", pady=(4, 0))
            tk.Label(frame_label, text=campo,
                     font=("Segoe UI", 8), bg=CINZA_CARD, fg=cor_label
                     ).pack(side="left")
            if campo == "carga-horaria":
                tk.Label(frame_label, text="  (calculado automaticamente)",
                         font=("Segoe UI", 7), bg=CINZA_CARD, fg=TEXTO_SUB
                         ).pack(side="left")

            var = tk.StringVar(value=str(val) if val is not None else "")
            entrada = tk.Entry(self._frame_campos, textvariable=var,
                               bg="#1e1e2e", fg=BRANCO,
                               insertbackground=BRANCO,
                               relief="flat", font=("Segoe UI", 9),
                               highlightthickness=1,
                               highlightbackground=VERMELHO if campo in entry["campos_faltantes"] else CINZA_BORDA,
                               highlightcolor=AZUL)
            entrada.pack(fill="x", pady=(0, 2))
            self._edits_vars[campo] = var

            # Quando hr-entrada-saida muda, recalcula carga-horaria automaticamente
            if campo == "hr-entrada-saida":
                def _ao_mudar_horario(*args, _var=var):
                    val_hr = _var.get().strip()
                    m = re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})", val_hr)
                    if m and "carga-horaria" in self._edits_vars:
                        try:
                            dt_e = datetime.strptime(m.group(1), "%H:%M")
                            dt_s = datetime.strptime(m.group(2), "%H:%M")
                            horas = int((dt_s - dt_e).total_seconds() / 3600)
                            self._edits_vars["carga-horaria"].set(str(horas))
                        except Exception:
                            pass
                var.trace_add("write", _ao_mudar_horario)

    def _aplicar_edicoes(self):
        if not self._linha_selecionada:
            return
        entry = self._linha_selecionada
        for campo, var in self._edits_vars.items():
            val = var.get().strip()
            if not val:
                entry["dados"][campo] = None
            elif campo == "carga-horaria":
                # Garante que carga-horaria seja salvo como inteiro
                try:
                    entry["dados"][campo] = int(val)
                except ValueError:
                    entry["dados"][campo] = val
            elif campo in ("CPF", "telefone", "ano/curso"):
                entry["dados"][campo] = formatar_numero(val)
            else:
                entry["dados"][campo] = val

            if val and campo in entry["campos_faltantes"]:
                entry["campos_faltantes"].remove(campo)

        # Recalcula carga-horaria a partir do hr-entrada-saida se não editado manualmente
        hr = entry["dados"].get("hr-entrada-saida", "")
        if hr:
            m = re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})", str(hr))
            if m and entry["dados"].get("carga-horaria") is None:
                try:
                    dt_e = datetime.strptime(m.group(1), "%H:%M")
                    dt_s = datetime.strptime(m.group(2), "%H:%M")
                    entry["dados"]["carga-horaria"] = int((dt_s - dt_e).total_seconds() / 3600)
                except Exception:
                    pass

        self._popular_tabela()
        messagebox.showinfo("OK", "Edições aplicadas!", parent=self)

    def _confirmar_salvar(self):
        incluidos = [e for e in self._resultados if e["incluir"] and not e["erro"]]
        if not incluidos:
            messagebox.showwarning("Atenção", "Nenhum arquivo marcado para salvar.", parent=self)
            return
        resp = messagebox.askyesno(
            "Confirmar",
            f"Salvar {len(incluidos)} contrato(s) no Excel?",
            parent=self
        )
        if resp:
            try:
                ok = salvar_resultados(self._resultados, self._pasta)
                rel = exportar_relatorio_erros(self._resultados, self._pasta)
                msg = f"{ok} contrato(s) salvos com sucesso!"
                if rel:
                    msg += f"\n\nRelatório de erros gerado em:\n{rel}"
                messagebox.showinfo("Sucesso", msg, parent=self)
                self._callback_salvo()
                self.destroy()
            except Exception as e:
                messagebox.showerror("Erro ao salvar", str(e), parent=self)


# ══════════════════════════════════════════════════════════════
#  JANELA DE CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════

class JanelaConfiguracoes(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Configurações")
        self.geometry("640x520")
        self.configure(bg=CINZA_BG)
        self.resizable(False, False)
        self.grab_set()
        self._cfg = carregar_config()
        self._construir_ui()

    def _construir_ui(self):
        tk.Label(self, text="⚙️  Configurações",
                 font=("Segoe UI", 13, "bold"), bg=CINZA_BG, fg=BRANCO
                 ).pack(padx=20, pady=(16, 8), anchor="w")

        # Excel
        frame_excel = tk.LabelFrame(self, text=" Local do Excel ",
                                    bg=CINZA_CARD, fg=TEXTO_SUB,
                                    font=("Segoe UI", 9),
                                    highlightthickness=0, bd=1,
                                    relief="groove")
        frame_excel.pack(fill="x", padx=20, pady=6)

        self._var_excel = tk.StringVar(value=self._cfg["caminho_excel"])
        tk.Entry(frame_excel, textvariable=self._var_excel,
                 bg=CINZA_BG, fg=BRANCO, insertbackground=BRANCO,
                 relief="flat", font=("Segoe UI", 9),
                 highlightthickness=1, highlightbackground=CINZA_BORDA
                 ).pack(side="left", fill="x", expand=True, padx=10, pady=8)

        tk.Button(frame_excel, text="📂",
                  command=self._escolher_excel,
                  bg=AZUL, fg="white", relief="flat",
                  font=("Segoe UI", 9), padx=8, cursor="hand2"
                  ).pack(side="right", padx=(0, 8), pady=8)

        # Padrões (regex)
        frame_padroes = tk.LabelFrame(self, text=" Expressões de extração (Regex) ",
                                      bg=CINZA_CARD, fg=TEXTO_SUB,
                                      font=("Segoe UI", 9),
                                      highlightthickness=0, bd=1, relief="groove")
        frame_padroes.pack(fill="both", expand=True, padx=20, pady=6)

        tk.Label(frame_padroes,
                 text="Edite os padrões de extração de texto dos PDFs:",
                 font=("Segoe UI", 8), bg=CINZA_CARD, fg=TEXTO_SUB
                 ).pack(padx=10, pady=(6, 2), anchor="w")

        canvas = tk.Canvas(frame_padroes, bg=CINZA_CARD, highlightthickness=0)
        scroll_y = ttk.Scrollbar(frame_padroes, orient="vertical", command=canvas.yview)
        self._frame_regex = tk.Frame(canvas, bg=CINZA_CARD)

        self._frame_regex.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._frame_regex, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set)
        scroll_y.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True, padx=8, pady=4)

        self._vars_regex = {}
        campos_regex = ["data-inicial", "data-final", "nome", "CPF",
                        "ano/curso", "supervisor", "hr-entrada-saida", "telefone"]
        for campo in campos_regex:
            tk.Label(self._frame_regex, text=campo,
                     font=("Segoe UI", 8, "bold"), bg=CINZA_CARD, fg=TEXTO_SUB
                     ).pack(anchor="w", padx=4, pady=(6, 0))
            var = tk.StringVar(value=self._cfg["padroes"].get(campo, ""))
            tk.Entry(self._frame_regex, textvariable=var,
                     bg=CINZA_BG, fg=BRANCO, insertbackground=BRANCO,
                     relief="flat", font=("Courier New", 8),
                     highlightthickness=1, highlightbackground=CINZA_BORDA,
                     width=70
                     ).pack(fill="x", padx=4, pady=(0, 2))
            self._vars_regex[campo] = var

        # carga-horaria: campo calculado, não tem regex própria
        tk.Label(self._frame_regex, text="carga-horaria",
                 font=("Segoe UI", 8, "bold"), bg=CINZA_CARD, fg=TEXTO_SUB
                 ).pack(anchor="w", padx=4, pady=(6, 0))
        tk.Label(self._frame_regex,
                 text="⚙  Calculado automaticamente a partir de hr-entrada-saida  (não possui regex)",
                 font=("Segoe UI", 8, "italic"), bg=CINZA_CARD, fg=TEXTO_SUB
                 ).pack(anchor="w", padx=4, pady=(0, 4))

        # Botões
        frame_bot = tk.Frame(self, bg=CINZA_BG)
        frame_bot.pack(fill="x", padx=20, pady=(6, 14))

        tk.Button(frame_bot, text="↩ Restaurar padrões",
                  command=self._restaurar,
                  bg=CINZA_CARD, fg=AMARELO, relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=12, pady=6, cursor="hand2"
                  ).pack(side="left")

        tk.Button(frame_bot, text="✖ Cancelar",
                  command=self.destroy,
                  bg=CINZA_CARD, fg=VERMELHO, relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=12, pady=6, cursor="hand2"
                  ).pack(side="right", padx=(6, 0))

        tk.Button(frame_bot, text="💾 Salvar configurações",
                  command=self._salvar,
                  bg=VERDE, fg="#0f172a", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=12, pady=6, cursor="hand2"
                  ).pack(side="right")

    def _escolher_excel(self):
        caminho = filedialog.asksaveasfilename(
            title="Escolha onde salvar o Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="estags.xlsx"
        )
        if caminho:
            self._var_excel.set(caminho)

    def _restaurar(self):
        resp = messagebox.askyesno("Confirmar", "Restaurar todos os padrões para o padrão original?", parent=self)
        if resp:
            for campo, var in self._vars_regex.items():
                var.set(PADROES_PADRAO.get(campo, ""))

    def _salvar(self):
        self._cfg["caminho_excel"] = self._var_excel.get().strip()
        for campo, var in self._vars_regex.items():
            self._cfg["padroes"][campo] = var.get().strip()
        salvar_config(self._cfg)
        messagebox.showinfo("Salvo", "Configurações salvas com sucesso!", parent=self)
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  JANELA DE HISTÓRICO
# ══════════════════════════════════════════════════════════════

class JanelaHistorico(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Histórico de execuções")
        self.geometry("640x360")
        self.configure(bg=CINZA_BG)
        self.resizable(True, False)
        self.grab_set()
        self._construir_ui()

    def _construir_ui(self):
        tk.Label(self, text="🕐  Histórico de execuções",
                 font=("Segoe UI", 12, "bold"), bg=CINZA_BG, fg=BRANCO
                 ).pack(padx=20, pady=(14, 8), anchor="w")

        frame = tk.Frame(self, bg=CINZA_CARD,
                         highlightthickness=1, highlightbackground=CINZA_BORDA)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 14))

        cols = ("data", "pasta", "salvos", "total")
        tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="none")
        tree.heading("data",   text="Data / Hora")
        tree.heading("pasta",  text="Pasta")
        tree.heading("salvos", text="Salvos")
        tree.heading("total",  text="Total PDFs")
        tree.column("data",   width=130, anchor="center")
        tree.column("pasta",  width=360)
        tree.column("salvos", width=60, anchor="center")
        tree.column("total",  width=80, anchor="center")

        style = ttk.Style(self)
        style.configure("Hist.Treeview",
                        background=CINZA_CARD, foreground=BRANCO,
                        fieldbackground=CINZA_CARD, rowheight=24,
                        font=("Segoe UI", 9))
        style.configure("Hist.Treeview.Heading",
                        background=CINZA_BG, foreground=TEXTO_SUB,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        tree.configure(style="Hist.Treeview")

        historico = carregar_historico()
        for item in reversed(historico):
            tree.insert("", "end", values=(
                item.get("data", ""),
                item.get("pasta", ""),
                item.get("salvos", ""),
                item.get("total", ""),
            ))

        if not historico:
            tree.insert("", "end", values=("—", "Nenhuma execução registrada", "—", "—"))

        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scroll.set)
        scroll.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)


# ══════════════════════════════════════════════════════════════
#  JANELA PRINCIPAL
# ══════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Processador de Contratos de Estágio")
        self.geometry("820x580")
        self.minsize(700, 480)
        self.resizable(True, True)
        self.configure(bg=CINZA_BG)
        self._pasta_selecionada = None
        self._em_processamento = False
        self._resultados_leitura = []
        self._construir_ui()
        self._configurar_drag_drop()

    # ── UI ──────────────────────────────────────────────────

    def _construir_ui(self):
        # Menu / barra superior
        frame_top = tk.Frame(self, bg=CINZA_BG)
        frame_top.pack(fill="x", padx=20, pady=(14, 4))

        tk.Label(frame_top, text="📄  Contratos de Estágio",
                 font=("Segoe UI", 15, "bold"), bg=CINZA_BG, fg=BRANCO
                 ).pack(side="left")

        tk.Button(frame_top, text="🕐 Histórico",
                  command=lambda: JanelaHistorico(self),
                  bg=CINZA_CARD, fg=TEXTO_SUB, relief="flat",
                  font=("Segoe UI", 8), padx=10, pady=4, cursor="hand2"
                  ).pack(side="right", padx=2)

        tk.Button(frame_top, text="⚙️ Config",
                  command=lambda: JanelaConfiguracoes(self),
                  bg=CINZA_CARD, fg=TEXTO_SUB, relief="flat",
                  font=("Segoe UI", 8), padx=10, pady=4, cursor="hand2"
                  ).pack(side="right", padx=2)

        # Linha pasta + arrastar
        frame_pasta = tk.Frame(self, bg=CINZA_CARD,
                               highlightthickness=1, highlightbackground=CINZA_BORDA)
        frame_pasta.pack(fill="x", padx=20, pady=6)

        self._lbl_drop = tk.Label(frame_pasta,
                                  text="📂  Arraste uma pasta aqui  ou  clique em Selecionar",
                                  font=("Segoe UI", 9), bg=CINZA_CARD, fg=TEXTO_SUB,
                                  padx=12, pady=8)
        self._lbl_drop.pack(side="left", fill="x", expand=True)

        tk.Button(frame_pasta, text="Selecionar pasta",
                  command=self._selecionar_pasta,
                  bg=AZUL, fg="white", relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=10, pady=4, cursor="hand2",
                  activebackground="#2563eb"
                  ).pack(side="right", padx=6, pady=6)

        # Excel info + botão abrir
        frame_excel = tk.Frame(self, bg=CINZA_BG)
        frame_excel.pack(fill="x", padx=20, pady=(0, 4))

        cfg = carregar_config()
        self._lbl_excel = tk.Label(frame_excel,
                                   text=f"📊 Excel: {cfg['caminho_excel']}",
                                   font=("Segoe UI", 7), bg=CINZA_BG, fg=TEXTO_SUB,
                                   anchor="w")
        self._lbl_excel.pack(side="left", fill="x", expand=True)

        self._btn_abrir_excel = tk.Button(frame_excel, text="Abrir Excel",
                  command=self._abrir_excel,
                  bg=CINZA_CARD, fg=VERDE, relief="flat",
                  font=("Segoe UI", 8, "bold"),
                  padx=8, pady=2, cursor="hand2"
                  )
        self._btn_abrir_excel.pack(side="right")

        # Tabela de arquivos
        frame_tabela = tk.Frame(self, bg=CINZA_CARD,
                                highlightthickness=1, highlightbackground=CINZA_BORDA)
        frame_tabela.pack(fill="both", expand=True, padx=20, pady=4)

        cols = ("arquivo", "status", "detalhe")
        self._tree = ttk.Treeview(frame_tabela, columns=cols, show="headings",
                                  selectmode="browse")
        self._tree.heading("arquivo", text="Arquivo")
        self._tree.heading("status",  text="Status")
        self._tree.heading("detalhe", text="Detalhe")
        self._tree.column("arquivo", width=360, minwidth=180)
        self._tree.column("status",  width=130, minwidth=80, anchor="center")
        self._tree.column("detalhe", width=280, minwidth=100)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview",
                        background=CINZA_CARD, foreground=BRANCO,
                        fieldbackground=CINZA_CARD, rowheight=26,
                        font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                        background=CINZA_BG, foreground=TEXTO_SUB,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#3b3b5c")])

        self._tree.tag_configure("ok",      foreground=VERDE)
        self._tree.tag_configure("erro",    foreground=VERMELHO)
        self._tree.tag_configure("pulado",  foreground=TEXTO_SUB)
        self._tree.tag_configure("lendo",   foreground=AMARELO)

        # Menu de contexto (botão direito)
        self._menu_contexto = tk.Menu(self, tearoff=0, bg=CINZA_CARD, fg=BRANCO,
                                      activebackground=AZUL, activeforeground=BRANCO,
                                      font=("Segoe UI", 9))
        self._menu_contexto.add_command(label="🔄 Reprocessar este arquivo",
                                        command=self._reprocessar_selecionado)
        self._tree.bind("<Button-3>", self._mostrar_menu_contexto)

        scroll = ttk.Scrollbar(frame_tabela, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscroll=scroll.set)
        scroll.pack(side="right", fill="y")
        self._tree.pack(fill="both", expand=True)

        # Progresso
        frame_prog = tk.Frame(self, bg=CINZA_BG)
        frame_prog.pack(fill="x", padx=20, pady=(4, 2))

        self._lbl_prog = tk.Label(frame_prog, text="",
                                  font=("Segoe UI", 8), bg=CINZA_BG, fg=TEXTO_SUB)
        self._lbl_prog.pack(side="right")

        style.configure("Custom.Horizontal.TProgressbar",
                        troughcolor=CINZA_CARD, background=VERDE,
                        borderwidth=0, thickness=7)
        self._progresso = ttk.Progressbar(frame_prog,
                                          style="Custom.Horizontal.TProgressbar",
                                          orient="horizontal", mode="determinate")
        self._progresso.pack(fill="x", expand=True, side="left", padx=(0, 8))

        # Rodapé botões
        frame_bot = tk.Frame(self, bg=CINZA_BG)
        frame_bot.pack(fill="x", padx=20, pady=(4, 14))

        self._lbl_resumo = tk.Label(frame_bot, text="",
                                    font=("Segoe UI", 9), bg=CINZA_BG, fg=TEXTO_SUB)
        self._lbl_resumo.pack(side="left")

        tk.Button(frame_bot, text="🗑 Limpar Excel",
                  command=self._limpar,
                  bg=CINZA_CARD, fg=VERMELHO, relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=12, pady=6, cursor="hand2"
                  ).pack(side="right", padx=(4, 0))

        tk.Button(frame_bot, text="🕑 Limpar histórico da pasta",
                  command=self._limpar_historico_pasta,
                  bg=CINZA_CARD, fg=AMARELO, relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=12, pady=6, cursor="hand2"
                  ).pack(side="right", padx=(4, 0))

        self._btn_processar = tk.Button(frame_bot, text="▶  Ler e pré-visualizar PDFs",
                  command=self._iniciar_leitura,
                  bg=VERDE, fg="#0f172a", relief="flat",
                  font=("Segoe UI", 9, "bold"),
                  padx=14, pady=6, cursor="hand2",
                  activebackground="#16a34a"
                  )
        self._btn_processar.pack(side="right")

    # ── Drag & Drop ─────────────────────────────────────────

    def _configurar_drag_drop(self):
        """Tenta ativar drag & drop via tkinterdnd2 se disponível."""
        try:
            from tkinterdnd2 import DND_FILES
            self.drop_target_register(DND_FILES)
            self.dnd_bind('<<Drop>>', self._ao_soltar_arquivo)
        except Exception:
            pass  # tkinterdnd2 não instalado; botão ainda funciona

    def _ao_soltar_arquivo(self, event):
        pasta = event.data.strip().strip("{}")
        if os.path.isdir(pasta):
            self._definir_pasta(pasta)

    # ── Ações ───────────────────────────────────────────────

    def _selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs")
        if pasta:
            self._definir_pasta(pasta)

    def _definir_pasta(self, pasta):
        self._pasta_selecionada = pasta
        self._lbl_drop.config(
            text=f"📁  {pasta}",
            fg=BRANCO
        )
        self._carregar_lista_arquivos(pasta)
        self._lbl_resumo.config(text="")
        self._progresso["value"] = 0
        self._lbl_prog.config(text="")
        # Atualiza label do excel
        cfg = carregar_config()
        self._lbl_excel.config(text=f"📊 Excel: {cfg['caminho_excel']}")

    def _carregar_lista_arquivos(self, pasta):
        for row in self._tree.get_children():
            self._tree.delete(row)

        ja_processados = carregar_processados(pasta)
        arquivos = sorted(f for f in os.listdir(pasta) if f.lower().endswith(".pdf"))

        cfg = carregar_config()
        existentes = carregar_contratos_existentes(cfg["caminho_excel"])

        novos = 0
        for arq in arquivos:
            nc = pegar_numero_contrato(arq)
            if arq in ja_processados:
                status = f"{ICONES['pulado']} Já processado"
                tag = "pulado"
            elif nc in existentes:
                status = "⏭ Duplicata"
                tag = "pulado"
            else:
                status = f"{ICONES['pendente']} Pendente"
                tag = "lendo"
                novos += 1
            self._tree.insert("", "end", iid=arq,
                              values=(arq, status, ""), tags=(tag,))

        total = len(arquivos)
        self._lbl_resumo.config(
            text=f"{total} arquivo(s) encontrado(s)  —  {novos} novo(s)"
        )

    def _iniciar_leitura(self):
        if self._em_processamento:
            return
        if not self._pasta_selecionada:
            messagebox.showwarning("Atenção", "Selecione uma pasta primeiro!")
            return
        arquivos = [f for f in os.listdir(self._pasta_selecionada) if f.lower().endswith(".pdf")]
        if not arquivos:
            messagebox.showinfo("Aviso", "Nenhum PDF encontrado.")
            return

        self._em_processamento = True
        self._btn_processar.config(state="disabled", text="⏳ Lendo PDFs…")
        self._progresso["value"] = 0
        self._lbl_resumo.config(text="Lendo arquivos…")

        cfg = carregar_config()

        t = threading.Thread(
            target=self._thread_leitura,
            args=(self._pasta_selecionada, cfg["padroes"]),
            daemon=True
        )
        t.start()

    def _thread_leitura(self, pasta, padroes):
        def cb_prog(atual, total, nome):
            self.after(0, lambda: self._atualizar_leitura(atual, total, nome))

        resultados = ler_pdfs_para_preview(pasta, padroes, cb_prog)
        self.after(0, lambda: self._leitura_concluida(resultados))

    def _atualizar_leitura(self, atual, total, nome):
        pct = int(atual / total * 100)
        self._progresso["value"] = pct
        self._lbl_prog.config(text=f"{atual}/{total}")
        if self._tree.exists(nome):
            self._tree.item(nome, values=(nome, f"{ICONES['processando']} Lendo…", ""))
            self._tree.see(nome)

    def _leitura_concluida(self, resultados):
        self._em_processamento = False
        self._btn_processar.config(state="normal", text="▶  Ler e pré-visualizar PDFs")
        self._resultados_leitura = resultados
        self._progresso["value"] = 100

        # Atualiza tabela com status de leitura
        for entry in resultados:
            arq = entry["arquivo"]
            if entry["erro"]:
                status = "❌ Erro na leitura"
                tag = "erro"
                detalhe = entry["erro"]
            elif entry["duplicata"]:
                status = "⏭ Duplicata"
                tag = "pulado"
                detalhe = "Já existe na planilha"
            elif entry["ja_processado"]:
                status = "⏭ Já processado"
                tag = "pulado"
                detalhe = ""
            elif entry["campos_faltantes"]:
                status = "⚠️ Campos faltando"
                tag = "lendo"
                detalhe = f"Falta: {', '.join(entry['campos_faltantes'])}"
            else:
                status = "✅ Lido com sucesso"
                tag = "ok"
                detalhe = ""
            if self._tree.exists(arq):
                self._tree.item(arq, values=(arq, status, detalhe), tags=(tag,))

        ok = sum(1 for e in resultados if not e["erro"] and not e["duplicata"] and not e["ja_processado"])
        erros = sum(1 for e in resultados if e["erro"])
        dup = sum(1 for e in resultados if e["duplicata"] or e["ja_processado"])
        self._lbl_resumo.config(
            fg=VERDE,
            text=f"✅ {ok} novos  |  ⏭ {dup} pulados  |  ❌ {erros} erros  — clique em Pré-visualizar para revisar"
        )

        # Abre pré-visualização automaticamente
        JanelaPreview(self, resultados, self._pasta_selecionada,
                      callback_salvo=lambda: self._carregar_lista_arquivos(self._pasta_selecionada))

    def _mostrar_menu_contexto(self, event):
        iid = self._tree.identify_row(event.y)
        if iid:
            self._tree.selection_set(iid)
            self._menu_contexto.post(event.x_root, event.y_root)

    def _reprocessar_selecionado(self):
        selecionados = self._tree.selection()
        if not selecionados:
            return
        arq = selecionados[0]
        if not self._pasta_selecionada:
            return

        cfg = carregar_config()
        caminho_pdf = os.path.join(self._pasta_selecionada, arq)
        if not os.path.exists(caminho_pdf):
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{caminho_pdf}")
            return

        self._tree.item(arq, values=(arq, f"{ICONES['processando']} Reprocessando…", ""), tags=("lendo",))

        def _tarefa():
            try:
                dados, faltantes = extrair_dados_pdf(caminho_pdf, cfg["padroes"])
                dados["vazio"] = pegar_numero_contrato(arq)
                entry = {
                    "arquivo": arq, "nc": dados["vazio"],
                    "dados": dados, "campos_faltantes": faltantes,
                    "duplicata": False, "ja_processado": False,
                    "erro": None, "incluir": True,
                }
                self.after(0, lambda: JanelaPreview(
                    self, [entry], self._pasta_selecionada,
                    callback_salvo=lambda: self._carregar_lista_arquivos(self._pasta_selecionada)
                ))
            except Exception as e:
                self.after(0, lambda: (
                    self._tree.item(arq, values=(arq, "❌ Erro", str(e)), tags=("erro",)),
                    messagebox.showerror("Erro", str(e))
                ))

        threading.Thread(target=_tarefa, daemon=True).start()

    def _limpar(self):
        if not self._pasta_selecionada:
            resp = messagebox.askyesno("Confirmar",
                "Deseja apagar todos os dados do Excel?\nEsta ação não pode ser desfeita.")
            limpar_hist = False
        else:
            from tkinter.simpledialog import Dialog
            # Pergunta customizada com 3 opções via messagebox encadeado
            resp = messagebox.askyesno("Confirmar",
                "Deseja apagar todos os dados do Excel?\nEsta ação não pode ser desfeita.")
            limpar_hist = False
            if resp:
                limpar_hist = messagebox.askyesno("Histórico da pasta",
                    "Deseja também limpar o histórico de arquivos já processados desta pasta?\n\n"
                    "Isso fará com que todos os PDFs apareçam como novos na próxima leitura.")

        if resp:
            try:
                limpar_excel()
                if limpar_hist and self._pasta_selecionada:
                    p = caminho_processados(self._pasta_selecionada)
                    if os.path.exists(p):
                        os.remove(p)
                cfg = carregar_config()
                msg = f"Excel limpo com sucesso."
                if limpar_hist:
                    msg += "\nHistórico da pasta também foi apagado."
                messagebox.showinfo("Sucesso", msg)
                if self._pasta_selecionada:
                    self._carregar_lista_arquivos(self._pasta_selecionada)
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    def _limpar_historico_pasta(self):
        if not self._pasta_selecionada:
            messagebox.showwarning("Atenção", "Selecione uma pasta primeiro!")
            return
        p = caminho_processados(self._pasta_selecionada)
        if not os.path.exists(p):
            messagebox.showinfo("Info", "Nenhum histórico encontrado para esta pasta.")
            return
        resp = messagebox.askyesno("Confirmar",
            "Limpar o histórico de arquivos já processados desta pasta?\n\n"
            "Os dados no Excel não serão apagados, mas os PDFs voltarão a aparecer como novos.")
        if resp:
            os.remove(p)
            self._carregar_lista_arquivos(self._pasta_selecionada)
            messagebox.showinfo("Sucesso", "Histórico da pasta limpo!\nTodos os PDFs agora aparecem como novos.")

    def _abrir_excel(self):
        cfg = carregar_config()
        caminho = cfg["caminho_excel"]
        if not os.path.exists(caminho):
            messagebox.showinfo("Aviso", "O arquivo Excel ainda não existe.\nProcesse algum PDF primeiro.")
            return
        try:
            os.startfile(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{e}")


# ══════════════════════════════════════════════════════════════
#  PONTO DE ENTRADA
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
