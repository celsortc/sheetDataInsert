import pdfplumber
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import tkinter as tk
from tkinter import messagebox, filedialog, Tk

def escolher_pasta():
    root = Tk()
    root.withdraw()
    pasta = filedialog.askdirectory(title="Selecione a pasta com PDFs")
    return pasta

def processar_pdfs():
    print('Processando PDFs...')
    ordem = [
        "data-inicial",
        "data-final",
        "vazio",
        "nome",
        "CPF",
        "hr-entrada-saida",
        "carga-horaria",
        "ano/curso",
        "supervisor",  # agora aparece depois da carga horaria
        "telefone"
    ]

    # Define a pasta que será aberta e lista os pdfs
    pasta = escolher_pasta()

    if not pasta:
        raise Exception("Nenhuma pasta selecionada.") 

    arquivos = [f for f in os.listdir(pasta) if f.endswith(".pdf")]

    arquivo = os.path.join(pasta, "estags.xlsx")
    print('arquivo aq: ',arquivo)

    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Estags"
        ws.append(ordem)

    # função que vai formatar o nome, permitir que capitalize o nome com excecão de palavras como de, do, dos..
    def formatar_nome(nome):
        if not nome:
            return nome

        palavras_minusculas = {"da", "de", "do", "das", "dos"}

        palavras = nome.lower().split()
        resultado = []

        for i, palavra in enumerate(palavras):
            if palavra in palavras_minusculas and i != 0:
                resultado.append(palavra)
            else:
                resultado.append(palavra.capitalize())

        return " ".join(resultado)

    def formatar_numeros(num):
        if not num:
            return num
        
        num = re.sub(r'\D', '', num)
        
        return int(num)

    def pegar_numero(nome_arquivo):
        numero = nome_arquivo.split(' - ')[0]
        return numero

    dados = {}

    for arquivo in arquivos:
        print("Processando:", arquivo)

        caminho_pdf = os.path.join(pasta, arquivo)

        nc = pegar_numero(arquivo)
        dados["vazio"] = nc

        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            for pagina in pdf.pages:
                texto += pagina.extract_text() + "\n"



        padroes = {
            "data-inicial": r"Vigência de:\s*(.*?)\sAté",
            "data-final": r"até\s*(.*)",
            "nome": r"Nome:\s*(.*?)\s+Código",
            "CPF": r"CPF/MF:\s*(.*)",
            "ano/curso": r"Regularmente Matriculado:\s*(\d+)",
            "supervisor": r"Supervisor:\s*(.*?)\sCargo",
            
        }

        

        for campo, padrao in padroes.items():
            match = re.search(padrao, texto, re.IGNORECASE)

            if match:
                dados[campo] = match.group(1).strip()
            else:
                dados[campo] = None



        match_horario = re.search(
            r"Horário das\s*(\d{2}:\d{2})\s*as\s*(\d{2}:\d{2})",
            texto,
            re.IGNORECASE
        )

        if match_horario:
            entrada = match_horario.group(1)
            saida = match_horario.group(2)

            # dados["horario-entrada"] = entrada
            # dados["horario-saida"] = saida
            dados["hr-entrada-saida"] = entrada + " - " + saida

            entrada_dt = datetime.strptime(entrada, "%H:%M")
            saida_dt = datetime.strptime(saida, "%H:%M")

            carga = (saida_dt - entrada_dt).total_seconds() / 3600
            dados["carga-horaria"] = int(carga)

        fones = re.findall(r"Fone:\s*([^\n]+)", texto)

        if len(fones) >= 3:
            dados["telefone"] = fones[2].strip()
        else:
            dados["telefone"] = None



        if dados.get("nome"):
            dados["nome"] = formatar_nome(dados["nome"])

        if dados.get("CPF") and dados.get("telefone") and dados.get("ano/curso"):
            dados["CPF"] = formatar_numeros(dados["CPF"])
            dados["telefone"] = formatar_numeros(dados["telefone"])
            dados["ano/curso"] = formatar_numeros(dados["ano/curso"])




        linha = [dados.get(campo) for campo in ordem]
        ws.append(linha)

    arquivo_excel = os.path.join(pasta, "estags.xlsx")
    wb.save(arquivo_excel)

def limpar_excel():
    pastaLimpar = escolher_pasta()

    if not pastaLimpar:
        messagebox.showwarning("Aviso", "Escolha uma pasta primeiro!")
        return

    arquivo = os.path.join(pastaLimpar, "estags.xlsx")

    if not os.path.exists(arquivo):
        messagebox.showinfo("Info", "Nenhum arquivo Excel encontrado.")
        return

    # recria o arquivo do zero
    wb = Workbook()
    ws = wb.active
    ws.title = "Estags"

    # seu cabeçalho
    ws.append([
        "data-inicial",
        "data-final",
        "vazio",
        "nome",
        "CPF",
        "hr-entrada-saida",
        "carga-horaria",
        "ano/curso",
        "supervisor",
        "telefone"
    ])

    wb.save(arquivo)

    messagebox.showinfo("Sucesso", "Dados apagados com sucesso!")



def clicar_botao():
    messagebox.showinfo("Atenção", "Selecione a pasta com os PDFs!")
    try:
        processar_pdfs()
        messagebox.showinfo("Sucesso", "PDFs processados com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

janela = tk.Tk()
janela.title("Automação de Estagiários")
janela.geometry("300x150")



botao = tk.Button(
    janela,
    text="Processar PDFs",
    command=clicar_botao,
    height=2,
    width=20
)
botao_limpar = tk.Button(
    janela,
    text="Limpar Excel",
    command=limpar_excel,
    height=2,
    width=20
)

botao_limpar.pack(pady=10)
botao.pack(pady=30)

janela.mainloop()

